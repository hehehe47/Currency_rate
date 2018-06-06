# from influxdb import InfluxDBClient
import time
# from src.currency import Currency

# client = InfluxDBClient(database='cr')
# client.write_points([{
#     'measurement':'dollar',
#     'tags':{'name':'abc'},
#     'fields':{'rate':'1313.1312112'}
# }])

# a = client.query("select rate from dollar ;")  # WHERE bname ='工商银行'
# for i in a['dollar']:
#     print(i['rate'])
#     #
#     # for j in i:
#     #     print(j['bname'],j['rate'])
# print(a['dollar'])



import requests
import bs4

# header = {
# 'Host':'mybank.nbcb.com.cn',
# 'Connection':'keep-alive',
# 'Upgrade-Insecure-Requests':'1',
# 'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.170 Safari/537.36',
# 'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
# 'Accept-Encoding':'gzip, deflate, br',
# 'Accept-Language':'zh-CN,zh;q=0.9,en;q=0.8',
# 'Cookie':'Hm_lvt_8b9480950a6cadd80a66f238d3e4542e=1528097970,1528097991'
# }
# r = requests.post(DOB[8]['url'], headers=HEADERS['兴业银行']['header'], data=HEADERS['兴业银行']['data'])
# print(r.text)
# b = bs4.BeautifulSoup(r.text,'html.parser')
# print()
# ind = b.select('td').index(b.select('td')[0])
# print(ind)
# for i,j in enumerate(b.select('td')):
# print(i,j)
# print(b)
# print(a)
#
# import xlsxwriter
#
# worksheet = xlsxwriter.Workbook('1.xlsx')
# worksheet.close()
# rate =None
# print([rate,'??'][rate])
# i = DOB[8]
# c = Currency(i['name'], i['rt_type'], i['url'], i['reg'], i['index'], i['header'])
# c.get_usd()
# import time
# # print(time.strftime('%Y-%M-%D %H:%m:%s',time.localtime(1528097970)))
# print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(1924768203)))
# print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(1528272892.100)))
LOG_FILE = '..\\log\\currency_rate.xlsx'

import openpyxl
w = openpyxl.load_workbook(LOG_FILE)
s = w.get_active_sheet()
print(s.max_column)